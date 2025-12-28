#!/usr/bin/env python3
"""
Improved extraction using openpyxl with better image-to-row mapping
This version attempts to properly map images to their Excel row positions
"""
import json
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from zipfile import ZipFile
import xml.etree.ElementTree as ET

def extract_with_openpyxl_improved(excel_path, output_dir):
    """
    Extract products and images with proper row-based image matching
    """
    print(f"Extracting data with openpyxl (Improved)\n")
    print(f"Excel file: {excel_path}")
    print(f"Output directory: {output_dir}\n")
    
    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    images_dir = output_path / "product_images_corrected"
    images_dir.mkdir(parents=True, exist_ok=True)
    
    # Load workbook
    print("=== Loading Excel file ===\n")
    workbook = load_workbook(excel_path, data_only=True, keep_links=False)
    worksheet = workbook.active
    print(f"Worksheet: {worksheet.title}")
    print(f"Max row: {worksheet.max_row}\n")
    
    # Step 1: Extract images from XLSX archive
    print("=== Step 1: Extracting images from XLSX archive ===\n")
    
    xlsx_images = {}
    image_files = []
    
    try:
        with ZipFile(excel_path, 'r') as zip_ref:
            # Get all image files
            all_files = zip_ref.namelist()
            image_files = [f for f in all_files if f.startswith('xl/media/') and 
                          any(f.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp'])]
            
            print(f"Found {len(image_files)} images in XLSX archive\n")
            
            # Read drawing relationships to map images to cells
            # Look for drawing files
            drawing_files = [f for f in all_files if 'xl/drawings/drawing' in f and f.endswith('.xml')]
            rels_files = [f for f in all_files if 'xl/drawings/_rels/drawing' in f and f.endswith('.rels')]
            
            print(f"Found {len(drawing_files)} drawing files")
            print(f"Found {len(rels_files)} relationship files\n")
            
            # Extract images
            for idx, img_path in enumerate(sorted(image_files)):
                try:
                    image_data = zip_ref.read(img_path)
                    img_filename = Path(img_path).name
                    img_ext = Path(img_path).suffix or '.jpeg'
                    
                    output_filename = f"image_{idx + 1}_{img_filename}{img_ext}"
                    output_path_img = images_dir / output_filename
                    
                    with open(output_path_img, 'wb') as f:
                        f.write(image_data)
                    
                    xlsx_images[img_path] = {
                        'index': idx,
                        'original_path': img_path,
                        'filename': output_filename,
                        'path': f"product_images_corrected/{output_filename}",
                        'size': len(image_data),
                        'excel_row': None  # Will be filled from worksheet images
                    }
                    
                    if idx < 5:
                        print(f"  Saved: {output_filename} ({len(image_data)/1024:.2f} KB)")
                except Exception as e:
                    print(f"  Warning: Could not extract {img_path}: {e}")
    except Exception as e:
        print(f"Error reading XLSX: {e}")
        return None
    
    print(f"\nExtracted {len(xlsx_images)} images\n")
    
    # Step 2: Get image positions from worksheet
    print("=== Step 2: Getting image positions from worksheet ===\n")
    
    row_to_image_map = {}
    
    # Access worksheet images
    if hasattr(worksheet, '_images') and worksheet._images:
        print(f"Found {len(worksheet._images)} images in worksheet._images\n")
        
        for idx, img_obj in enumerate(worksheet._images):
            row = None
            col = None
            
            # Try to get anchor information
            try:
                anchor = img_obj.anchor
                if hasattr(anchor, '_from'):
                    # TwoCellAnchor
                    row = anchor._from.row + 1  # Convert to 1-indexed
                    col = anchor._from.col + 1
                elif hasattr(anchor, 'row'):
                    # OneCellAnchor
                    row = anchor.row + 1
                    col = anchor.col + 1
                
                if row:
                    # Map row to image
                    # The image index in worksheet._images corresponds to the order in xlsx_images
                    image_list = sorted(xlsx_images.keys())
                    if idx < len(image_list):
                        image_path = image_list[idx]
                        # Store mapping: row -> image data
                        row_to_image_map[row] = xlsx_images[image_path]
                        xlsx_images[image_path]['excel_row'] = row
                        
                        if idx < 10:
                            print(f"  Image {idx + 1}: Row {row}, Column {col or 'N/A'}")
            except Exception as e:
                if idx < 5:
                    print(f"  Warning: Could not get position for image {idx + 1}: {e}")
    
    print(f"\nMapped {len(row_to_image_map)} images to rows\n")
    
    # Step 3: Extract product data
    print("=== Step 3: Extracting product data ===\n")
    
    products = []
    sequential_image_index = 0
    image_list_sorted = sorted(xlsx_images.values(), key=lambda x: x['index'])
    
    # Extract products starting from row 3 (skip first 2 header rows)
    # Images start from row 8, but products might start earlier
    for row_num in range(1, worksheet.max_row + 1):
        row = worksheet[row_num]
        
        # Skip first 2 rows (definitely headers)
        if row_num <= 2:
            continue
        
        # In openpyxl, row indexing starts from 1: row[1] = column A, row[2] = column B, etc.
        # Get product code from column A (row[1]) and name from column B (row[2])
        product_code = ''
        product_name = ''
        
        if len(row) >= 1 and row[1].value:
            product_code = str(row[1].value).strip()
        if len(row) >= 2 and row[2].value:
            product_name = str(row[2].value).strip()
        
        # Skip if no product name
        if not product_name or product_name == '':
            continue
        
        # Skip only obvious header rows (very strict - only exact matches)
        if product_code == '序号' or product_name == '名称' or product_name == '产品图':
            continue
        
        # Extract product data (columns: A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12)
        battery_model = str(row[3].value).strip() if len(row) >= 3 and row[3].value else '无'
        unit = str(row[4].value).strip() if len(row) >= 4 and row[4].value else ''
        cost_price = float(row[5].value) if len(row) >= 5 and row[5].value else 0
        quantity = int(row[6].value) if len(row) >= 6 and row[6].value else 0
        specification = str(row[7].value).strip() if len(row) >= 7 and row[7].value else ''
        amount = float(row[8].value) if len(row) >= 8 and row[8].value else 0
        suggested_price = float(row[9].value) if len(row) >= 9 and row[9].value else 0
        market_value = float(row[10].value) if len(row) >= 10 and row[10].value else 0
        barcode = str(row[11].value).strip() if len(row) >= 11 and row[11].value else ''
        weight = float(row[12].value) if len(row) >= 12 and row[12].value else 0
        
        # Match image to this row
        matched_image = None
        
        # Method 1: Use row-based mapping if available (best method)
        if row_num in row_to_image_map:
            matched_image = row_to_image_map[row_num]
        # Method 2: Check nearby rows (images might be slightly offset)
        elif (row_num + 1) in row_to_image_map:
            matched_image = row_to_image_map[row_num + 1]
        elif (row_num - 1) in row_to_image_map:
            matched_image = row_to_image_map[row_num - 1]
        # Method 3: Use sequential index (fallback)
        elif sequential_image_index < len(image_list_sorted):
            matched_image = image_list_sorted[sequential_image_index]
            sequential_image_index += 1
        
        # Build image array
        matched_images = []
        if matched_image:
            matched_images.append({
                'filename': matched_image['filename'],
                'path': matched_image['path'],
                'size': matched_image['size']
            })
        
        product = {
            'rowNumber': row_num,
            '序号': product_code,
            '名称': product_name,
            '电池型号': battery_model,
            '单位': unit,
            '价格': cost_price,
            '数量': quantity,
            '规格': specification,
            '金额': amount,
            '建议价': suggested_price,
            '市值': market_value,
            '条码': barcode,
            '重量KG': weight,
            '图片': matched_images
        }
        
        products.append(product)
        
        if len(products) <= 5:
            img_info = matched_images[0]['filename'] if matched_images else 'None'
            match_method = 'row' if row_num in row_to_image_map else 'index'
            print(f"  Product {len(products)}: \"{product_name[:40]}...\" - Row {row_num}, Image: {img_info} ({match_method})")
    
    print(f"\nExtracted {len(products)} products\n")
    
    # Step 4: Save results
    print("=== Step 4: Saving results ===\n")
    
    json_path = output_path / "extracted_products_with_images.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    
    print(f"Products data saved to: {json_path}\n")
    
    # Summary
    products_with_images = sum(1 for p in products if p['图片'] and len(p['图片']) > 0)
    row_matched = sum(1 for p in products if p['rowNumber'] in row_to_image_map)
    
    print("=== Extraction Summary ===")
    print(f"Total products: {len(products)}")
    print(f"Total images: {len(xlsx_images)}")
    print(f"Products with images: {products_with_images}")
    print(f"Images matched by row position: {row_matched}")
    print(f"Images matched by sequential index: {products_with_images - row_matched}")
    print(f"\nExtraction complete!")
    
    return {
        'products': products,
        'images': list(xlsx_images.values()),
        'output_path': str(json_path)
    }

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python extract_with_openpyxl_improved.py <excel-file> [output-dir]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '../extracted_data_openpyxl'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        extract_with_openpyxl_improved(excel_path, output_dir)
    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

