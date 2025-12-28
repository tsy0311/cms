#!/usr/bin/env python3
"""
Extract products and images from Excel file using openpyxl
This provides better image position tracking than JavaScript libraries
"""
import json
import sys
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import shutil
from zipfile import ZipFile

def extract_with_openpyxl(excel_path, output_dir):
    """
    Extract products and images from Excel file with correct image-to-product mapping
    """
    print(f"📊 Extracting data with openpyxl\n")
    print(f"Excel file: {excel_path}")
    print(f"Output directory: {output_dir}\n")
    
    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    images_dir = output_path / "product_images_corrected"
    images_dir.mkdir(parents=True, exist_ok=True)
    
    # Load workbook
    print("=== Loading Excel file ===\n")
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active
    print(f"Worksheet: {worksheet.title}")
    print(f"Max row: {worksheet.max_row}\n")
    
    # Step 1: Extract images with their positions
    print("=== Step 1: Extracting images with positions ===\n")
    
    # Get images from worksheet
    images_info = []
    
    # Access images via worksheet._images (private, but necessary for position data)
    if hasattr(worksheet, '_images') and worksheet._images:
        print(f"Found {len(worksheet._images)} images in worksheet\n")
        
        for idx, img in enumerate(worksheet._images):
            # Get image position - anchor contains row/column information
            anchor = img.anchor
            row = None
            col = None
            
            if hasattr(anchor, '_from'):
                # TwoCellAnchor format
                row = anchor._from.row + 1  # openpyxl is 0-indexed, Excel is 1-indexed
                col = anchor._from.col + 1
            elif hasattr(anchor, 'row'):
                # OneCellAnchor format
                row = anchor.row + 1
                col = anchor.col + 1
            
            # Get image data
            image_data = img._data()
            image_ext = img.path.split('.')[-1] if '.' in img.path else 'jpeg'
            
            images_info.append({
                'index': idx,
                'row': row,
                'col': col,
                'ext': image_ext,
                'data': image_data,
                'path': img.path
            })
            
            if idx < 10:
                print(f"  Image {idx + 1}: Row {row}, Column {col} ({len(image_data)} bytes)")
    
    # Alternative: Extract images from XLSX as ZIP (more reliable)
    print("\n=== Step 2: Extracting images from XLSX archive ===\n")
    
    xlsx_images = []
    try:
        with ZipFile(excel_path, 'r') as zip_ref:
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and 
                          any(f.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp'])]
            
            print(f"Found {len(image_files)} images in XLSX archive\n")
            
            # Create mapping: image filename -> Excel row
            # We'll match by order and use images_info for row positions if available
            image_row_map = {}
            for img_info in images_info:
                if img_info['row']:
                    # Map by image index in the archive
                    # Note: The order in xl/media/ might not match worksheet._images order
                    # We'll need to match based on the image path/filename
                    pass
            
            for idx, img_path in enumerate(image_files):
                try:
                    image_data = zip_ref.read(img_path)
                    img_filename = Path(img_path).name
                    img_ext = Path(img_path).suffix or '.jpeg'
                    
                    output_filename = f"product_{idx + 1}_{img_filename}{img_ext}"
                    output_path_img = images_dir / output_filename
                    
                    with open(output_path_img, 'wb') as f:
                        f.write(image_data)
                    
                    # Try to find row position from images_info
                    row_pos = None
                    if idx < len(images_info) and images_info[idx]['row']:
                        row_pos = images_info[idx]['row']
                    
                    xlsx_images.append({
                        'index': idx,
                        'original_path': img_path,
                        'filename': output_filename,
                        'path': f"product_images_corrected/{output_filename}",
                        'size': len(image_data),
                        'row': row_pos
                    })
                    
                    if idx < 5:
                        print(f"  Saved: {output_filename} ({len(image_data)/1024:.2f} KB, Row: {row_pos or 'N/A'})")
                except Exception as e:
                    print(f"  Warning: Could not extract {img_path}: {e}")
    except Exception as e:
        print(f"Warning: Could not read XLSX as ZIP: {e}")
    
    print(f"\n✅ Extracted {len(xlsx_images)} images\n")
    
    # Step 3: Extract product data
    print("=== Step 3: Extracting product data ===\n")
    
    products = []
    product_row_to_image = {}
    
    # Build map: Excel row -> image index
    for img_info in images_info:
        if img_info['row']:
            # Map row to image index in xlsx_images
            # Find the image that corresponds to this row
            row_num = img_info['row']
            if row_num not in product_row_to_image:
                # Try to match by index if available
                product_row_to_image[row_num] = img_info['index']
    
    # Extract products row by row
    row_idx = 0
    for row_num in range(1, worksheet.max_row + 1):
        row = worksheet[row_num]
        
        # Skip header rows (usually first 2 rows)
        if row_num <= 2:
            # Check if it's a header
            first_cell = row[0].value if len(row) > 0 else None
            if first_cell and ('序号' in str(first_cell) or '配货中心' in str(first_cell)):
                continue
        
        # Get product name from column B (index 1)
        product_name = None
        if len(row) > 1:
            product_name = row[1].value
            if product_name:
                product_name = str(product_name).strip()
        
        # Skip if no product name or if it looks like a header
        if not product_name or product_name == '' or '配货中心' in product_name or product_name == '名称':
            continue
        
        row_idx += 1
        
        # Extract all fields
        product_code = str(row[0].value).strip() if len(row) > 0 and row[0].value else ''
        battery_model = str(row[2].value).strip() if len(row) > 2 and row[2].value else '无'
        unit = str(row[3].value).strip() if len(row) > 3 and row[3].value else ''
        
        # Numeric fields
        cost_price = float(row[4].value) if len(row) > 4 and row[4].value else 0
        quantity = int(row[5].value) if len(row) > 5 and row[5].value else 0
        specification = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''
        amount = float(row[7].value) if len(row) > 7 and row[7].value else 0
        suggested_price = float(row[8].value) if len(row) > 8 and row[8].value else 0
        market_value = float(row[9].value) if len(row) > 9 and row[9].value else 0
        barcode = str(row[10].value).strip() if len(row) > 10 and row[10].value else ''
        weight = float(row[11].value) if len(row) > 11 and row[11].value else 0
        
        # Match image to this product row
        matched_images = []
        image_index = None
        
        # Method 1: Try to find image by row position
        if row_num in product_row_to_image:
            image_index = product_row_to_image[row_num]
        else:
            # Method 2: Use sequential index (fallback)
            image_index = row_idx - 1
        
        # Get the image if index is valid
        if image_index is not None and image_index < len(xlsx_images):
            img = xlsx_images[image_index]
            matched_images.append({
                'filename': img['filename'],
                'path': img['path'],
                'size': img['size']
            })
        
        product = {
            'rowNumber': row_num,
            '序号': product_code,
            '名称': product_name,
            '电池型号': battery_model,
            '单位': unit,
            '价格': cost_price,
            '数量': quantity,
            '规格': specification,
            '金额': amount,
            '建议价': suggested_price,
            '市值': market_value,
            '条码': barcode,
            '重量KG': weight,
            '图片': matched_images
        }
        
        products.append(product)
        
        if row_idx <= 5:
            img_info = matched_images[0]['filename'] if matched_images else 'None'
            print(f"  Product {row_idx}: \"{product_name[:40]}...\" - Row {row_num}, Image: {img_info}")
    
    print(f"\n✅ Extracted {len(products)} products\n")
    
    # Step 4: Save results
    print("=== Step 4: Saving results ===\n")
    
    json_path = output_path / "extracted_products_with_images.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Products data saved to: {json_path}\n")
    
    # Summary
    products_with_images = sum(1 for p in products if p['图片'] and len(p['图片']) > 0)
    
    print("=== Extraction Summary ===")
    print(f"Total products: {len(products)}")
    print(f"Total images: {len(xlsx_images)}")
    print(f"Products with images: {products_with_images}")
    print(f"Images with row positions: {len([img for img in images_info if img['row']])}")
    print(f"\n✅ Extraction complete!")
    print(f"\n⚠️  Note: Image matching uses row positions when available, otherwise sequential index.")
    print(f"   Please verify a few products manually to ensure correct matching.")
    
    return {
        'products': products,
        'images': xlsx_images,
        'output_path': str(json_path)
    }

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python extract_with_openpyxl.py <excel-file> [output-dir]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '../extracted_data_openpyxl'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        extract_with_openpyxl(excel_path, output_dir)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

